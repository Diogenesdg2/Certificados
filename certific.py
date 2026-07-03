import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import json
import os
import sys
import smtplib
import ssl
import threading
import schedule
import time
import base64
import hashlib
import sqlite3
from datetime import datetime, date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
try:
    import pystray
    from PIL import Image, ImageDraw
    TRAY_DISPONIVEL = True
except ImportError:
    TRAY_DISPONIVEL = False

import winreg

STARTUP_REG_KEY  = r"Software\Microsoft\Windows\CurrentVersion\Run"
STARTUP_APP_NAME = "GerenciadorCertificados"

def _startup_habilitado() -> bool:
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, STARTUP_REG_KEY, 0, winreg.KEY_READ)
        winreg.QueryValueEx(key, STARTUP_APP_NAME)
        winreg.CloseKey(key)
        return True
    except FileNotFoundError:
        return False

def _habilitar_startup():
    caminho = os.path.abspath(sys.argv[0])
    if caminho.endswith(".py"):
        exe = os.path.join(os.path.dirname(sys.executable), "pythonw.exe")
        if not os.path.exists(exe):
            exe = sys.executable
        valor = f'"{exe}" "{caminho}"'
    else:
        valor = f'"{caminho}"'
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, STARTUP_REG_KEY, 0, winreg.KEY_SET_VALUE)
    winreg.SetValueEx(key, STARTUP_APP_NAME, 0, winreg.REG_SZ, valor)
    winreg.CloseKey(key)

def _desabilitar_startup():
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, STARTUP_REG_KEY, 0, winreg.KEY_SET_VALUE)
        winreg.DeleteValue(key, STARTUP_APP_NAME)
        winreg.CloseKey(key)
    except FileNotFoundError:
        pass

from cryptography.hazmat.primitives.serialization.pkcs12 import load_key_and_certificates
from cryptography.x509 import load_pem_x509_certificate
from cryptography.hazmat.backends import default_backend
from cryptography.fernet import Fernet

# ─────────────────────────────────────────────
#  BANCO DE DADOS SQLite
# ─────────────────────────────────────────────
def _pasta_base() -> str:
    """Retorna sempre a pasta do executável/script, independente do diretório atual."""
    if getattr(sys, 'frozen', False):
        # Rodando como .exe (PyInstaller)
        return os.path.dirname(sys.executable)
    else:
        # Rodando como .py
        return os.path.dirname(os.path.abspath(__file__))

DB_FILE    = os.path.join(_pasta_base(), "certificados.db")
CHAVE_FILE = os.path.join(_pasta_base(), "chave.key")

_db_lock = threading.Lock()

def _conectar():
    conn = sqlite3.connect(DB_FILE, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=DELETE") # modo padrao, seguro em rede compartilhada
    conn.execute("PRAGMA busy_timeout=30000")  # aguarda ate 30s se banco estiver ocupado
    conn.execute("PRAGMA synchronous=FULL")    # maxima seguranca para evitar corrupcao
    conn.execute("PRAGMA foreign_keys=ON")
    return conn

def _executar_com_retry(func, tentativas=5, espera=0.5):
    """Executa func() com retry automatico em caso de database is locked."""
    import time as _time
    ultimo_erro = None
    for i in range(tentativas):
        try:
            return func()
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower():
                ultimo_erro = e
                _time.sleep(espera * (i + 1))
            else:
                raise
    raise ultimo_erro

def _init_db():
    """Cria as tabelas se nao existirem."""
    def _fazer():
        with _conectar() as conn:
            return conn
    _executar_com_retry(_fazer)  # testa conexao antes de criar tabelas
    with _db_lock, _conectar() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS certificados (
                id            TEXT PRIMARY KEY,
                tipo          TEXT NOT NULL DEFAULT 'A1',
                nome          TEXT NOT NULL,
                responsavel   TEXT DEFAULT '',
                vencimento    TEXT NOT NULL,
                obs           TEXT DEFAULT '',
                emails        TEXT DEFAULT '',
                arquivo_nome  TEXT DEFAULT '',
                arquivo_b64   TEXT DEFAULT '',
                arquivo_ext   TEXT DEFAULT 'pfx',
                senha_enc     TEXT DEFAULT '',
                ultimo_alerta TEXT DEFAULT '',
                dias_alerta   INTEGER DEFAULT 15,
                enviar_alerta INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS historico (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                cert_id   TEXT NOT NULL,
                data      TEXT NOT NULL,
                hora      TEXT NOT NULL,
                acao      TEXT NOT NULL,
                info_json TEXT DEFAULT '{}'
            );
            CREATE TABLE IF NOT EXISTS configuracoes (
                chave TEXT PRIMARY KEY,
                valor TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS senha_mestre (
                id   INTEGER PRIMARY KEY CHECK (id = 1),
                hash TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS log_emails (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                data_hora     TEXT NOT NULL,
                cert_id       TEXT NOT NULL,
                cert_nome     TEXT NOT NULL,
                cert_tipo     TEXT NOT NULL,
                destinatarios TEXT NOT NULL,
                assunto       TEXT NOT NULL,
                status        TEXT NOT NULL,
                erro          TEXT DEFAULT '',
                origem        TEXT NOT NULL DEFAULT 'automatico',
                lido          TEXT DEFAULT 'Pendente',
                data_leitura  TEXT DEFAULT ''
            );
            -- Migração: adiciona colunas se não existirem (seguro para bancos existentes)
            PRAGMA user_version;
        """)

# ─────────────────────────────────────────────
#  CORES DO TEMA
# ─────────────────────────────────────────────
COR_PRIMARIA   = "#1a2a4a"
COR_SECUNDARIA = "#2563eb"
COR_ACENTO     = "#38bdf8"
COR_BG         = "#f0f4f8"
COR_BG_TABLE   = "#ffffff"
COR_TEXTO_CLR  = "#ffffff"
COR_OK         = "#dcfce7"
COR_ATENCAO    = "#fef9c3"
COR_CRITICO    = "#ffedd5"
COR_VENCIDO    = "#fee2e2"
COR_OK_FG      = "#166534"
COR_ATENCAO_FG = "#854d0e"
COR_CRITICO_FG = "#9a3412"
COR_VENCIDO_FG = "#991b1b"

# ─────────────────────────────────────────────
#  CRIPTOGRAFIA DE SENHA
# ─────────────────────────────────────────────
def _obter_fernet():
    if os.path.exists(CHAVE_FILE):
        with open(CHAVE_FILE, "rb") as f:
            chave = f.read()
    else:
        chave = Fernet.generate_key()
        with open(CHAVE_FILE, "wb") as f:
            f.write(chave)
    return Fernet(chave)

def criptografar_senha(senha: str) -> str:
    if not senha:
        return ""
    return _obter_fernet().encrypt(senha.encode()).decode()

def descriptografar_senha(senha_enc: str) -> str:
    if not senha_enc:
        return ""
    try:
        return _obter_fernet().decrypt(senha_enc.encode()).decode()
    except Exception:
        return ""

DIAS_ALERTA_INICIO = 30

# ─────────────────────────────────────────────
#  SENHA MESTRE
# ─────────────────────────────────────────────
def _hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode()).hexdigest()

def senha_mestre_definida() -> bool:
    with _db_lock, _conectar() as conn:
        row = conn.execute("SELECT hash FROM senha_mestre WHERE id=1").fetchone()
        return row is not None

def verificar_senha_mestre(senha: str) -> bool:
    with _db_lock, _conectar() as conn:
        row = conn.execute("SELECT hash FROM senha_mestre WHERE id=1").fetchone()
        if not row:
            return False
        return row["hash"] == _hash_senha(senha)

def definir_senha_mestre(senha: str):
    with _db_lock, _conectar() as conn:
        conn.execute(
            "INSERT INTO senha_mestre (id, hash) VALUES (1, ?) "
            "ON CONFLICT(id) DO UPDATE SET hash=excluded.hash",
            (_hash_senha(senha),)
        )

def pedir_senha_mestre(parent=None) -> bool:
    if not senha_mestre_definida():
        win = tk.Toplevel(parent)
        win.title("Definir Senha Mestre")
        win.resizable(False, False)
        win.grab_set()
        win.geometry("360x220")

        hdr = tk.Frame(win, bg=COR_PRIMARIA, height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  Definir Senha Mestre",
                 bg=COR_PRIMARIA, fg="#ffffff",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=12, pady=10)

        ttk.Label(win, text="Crie uma senha mestre para proteger\na visualizacao das senhas dos certificados.",
                  font=("Segoe UI", 9), justify="center").pack(pady=(14, 6))

        f = ttk.Frame(win)
        f.pack(pady=4)
        ttk.Label(f, text="Nova senha:").grid(row=0, column=0, sticky="e", padx=6, pady=4)
        v1 = tk.StringVar()
        ttk.Entry(f, textvariable=v1, show="*", width=22).grid(row=0, column=1, pady=4)
        ttk.Label(f, text="Confirmar:").grid(row=1, column=0, sticky="e", padx=6, pady=4)
        v2 = tk.StringVar()
        ttk.Entry(f, textvariable=v2, show="*", width=22).grid(row=1, column=1, pady=4)

        resultado = [False]
        def _confirmar():
            s1, s2 = v1.get().strip(), v2.get().strip()
            if not s1:
                messagebox.showwarning("Atencao", "Digite uma senha.", parent=win)
                return
            if s1 != s2:
                messagebox.showerror("Erro", "As senhas nao conferem.", parent=win)
                return
            definir_senha_mestre(s1)
            resultado[0] = True
            win.destroy()

        bf = ttk.Frame(win)
        bf.pack(pady=8)
        ttk.Button(bf, text="Definir", command=_confirmar).pack(side="left", padx=5)
        ttk.Button(bf, text="Cancelar", command=win.destroy).pack(side="left", padx=5)
        win.wait_window()
        return resultado[0]
    else:
        senha = simpledialog.askstring(
            "Senha Mestre", "Digite a senha mestre:",
            show="*", parent=parent
        )
        if senha is None:
            return False
        if not verificar_senha_mestre(senha):
            messagebox.showerror("Erro", "Senha mestre incorreta.", parent=parent)
            return False
        return True

# ─────────────────────────────────────────────
#  ARQUIVO DE CERTIFICADO
# ─────────────────────────────────────────────
def arquivo_para_base64(caminho: str) -> str:
    with open(caminho, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def base64_para_arquivo(b64: str, destino: str):
    with open(destino, "wb") as f:
        f.write(base64.b64decode(b64))

TEMPLATE_PADRAO = {
    "assunto": "[Certificado Digital] {nome} - vence em {dias} dia(s)",
    "corpo": """<html><body style="font-family:Arial,sans-serif;padding:20px">
<div style="border-left:6px solid {cor};padding:15px;background:#fafafa;border-radius:4px">
  <h2 style="color:{cor};margin:0 0 10px">Alerta de Certificado Digital</h2>
  <table style="border-collapse:collapse;width:100%">
    <tr><td style="padding:6px;font-weight:bold;width:160px">Nome/Titular:</td>
        <td style="padding:6px">{nome}</td></tr>
    <tr style="background:#f0f0f0">
        <td style="padding:6px;font-weight:bold">Tipo:</td>
        <td style="padding:6px">{tipo}</td></tr>
    <tr><td style="padding:6px;font-weight:bold">Responsavel:</td>
        <td style="padding:6px">{responsavel}</td></tr>
    <tr style="background:#f0f0f0">
        <td style="padding:6px;font-weight:bold">Vencimento:</td>
        <td style="padding:6px;color:{cor};font-weight:bold">{vencimento} - {situacao}</td></tr>
    <tr><td style="padding:6px;font-weight:bold">Observacao:</td>
        <td style="padding:6px">{obs}</td></tr>
  </table>
</div>
<p style="color:#777;font-size:12px;margin-top:20px">Mensagem automatica - Gerenciador de Certificados Digitais</p>
</body></html>"""
}

# ─────────────────────────────────────────────
#  PERSISTENCIA — CERTIFICADOS
# ─────────────────────────────────────────────
def _row_to_cert(row) -> dict:
    c = dict(row)
    c["historico"] = []
    return c

def carregar_certificados() -> list:
    def _fazer():
        with _db_lock, _conectar() as conn:
            rows = conn.execute("SELECT * FROM certificados ORDER BY vencimento").fetchall()
            certs = []
            for row in rows:
                c = _row_to_cert(row)
                hist_rows = conn.execute(
                    "SELECT * FROM historico WHERE cert_id=? ORDER BY id DESC LIMIT 90",
                    (c["id"],)
                ).fetchall()
                c["historico"] = [
                    {**json.loads(h["info_json"]),
                     "data": h["data"], "hora": h["hora"], "acao": h["acao"]}
                    for h in reversed(hist_rows)
                ]
                certs.append(c)
            return certs
    return _executar_com_retry(_fazer)

def _carregar_certificados_original() -> list:
    with _db_lock, _conectar() as conn:
        rows = conn.execute("SELECT * FROM certificados ORDER BY vencimento").fetchall()
        certs = []
        for row in rows:
            c = _row_to_cert(row)
            hist_rows = conn.execute(
                "SELECT * FROM historico WHERE cert_id=? ORDER BY id DESC LIMIT 90",
                (c["id"],)
            ).fetchall()
            c["historico"] = [
                {**json.loads(h["info_json"]),
                 "data": h["data"], "hora": h["hora"], "acao": h["acao"]}
                for h in reversed(hist_rows)
            ]
            certs.append(c)
    return certs

def salvar_certificado(cert: dict):
    def _fazer():
      with _db_lock, _conectar() as conn:
        conn.execute("""
            INSERT INTO certificados
                (id, tipo, nome, responsavel, vencimento, obs, emails,
                 arquivo_nome, arquivo_b64, arquivo_ext, senha_enc, ultimo_alerta,
                 dias_alerta, enviar_alerta)
            VALUES
                (:id,:tipo,:nome,:responsavel,:vencimento,:obs,:emails,
                 :arquivo_nome,:arquivo_b64,:arquivo_ext,:senha_enc,:ultimo_alerta,
                 :dias_alerta,:enviar_alerta)
            ON CONFLICT(id) DO UPDATE SET
                tipo=excluded.tipo, nome=excluded.nome,
                responsavel=excluded.responsavel, vencimento=excluded.vencimento,
                obs=excluded.obs, emails=excluded.emails,
                arquivo_nome=excluded.arquivo_nome, arquivo_b64=excluded.arquivo_b64,
                arquivo_ext=excluded.arquivo_ext, senha_enc=excluded.senha_enc,
                ultimo_alerta=excluded.ultimo_alerta,
                dias_alerta=excluded.dias_alerta,
                enviar_alerta=excluded.enviar_alerta
        """, {
            "id":            cert.get("id", ""),
            "tipo":          cert.get("tipo", "A1"),
            "nome":          cert.get("nome", ""),
            "responsavel":   cert.get("responsavel", ""),
            "vencimento":    cert.get("vencimento", ""),
            "obs":           cert.get("obs", ""),
            "emails":        cert.get("emails", ""),
            "arquivo_nome":  cert.get("arquivo_nome", ""),
            "arquivo_b64":   cert.get("arquivo_b64", ""),
            "arquivo_ext":   cert.get("arquivo_ext", "pfx"),
            "senha_enc":     cert.get("senha_enc", ""),
            "ultimo_alerta": cert.get("ultimo_alerta", ""),
            "dias_alerta":   int(cert.get("dias_alerta", 15)),
            "enviar_alerta": 1 if cert.get("enviar_alerta", True) else 0,
          })
    _executar_com_retry(_fazer)

def excluir_certificado_db(cert_id: str):
    def _fazer():
        with _db_lock, _conectar() as conn:
            conn.execute("DELETE FROM historico WHERE cert_id=?", (cert_id,))
            conn.execute("DELETE FROM certificados WHERE id=?", (cert_id,))
    _executar_com_retry(_fazer)

def registrar_historico_db(cert_id: str, acao: str, **extras):
    info = json.dumps(extras, ensure_ascii=False)
    def _fazer():
        with _db_lock, _conectar() as conn:
            conn.execute(
                "INSERT INTO historico (cert_id, data, hora, acao, info_json) VALUES (?,?,?,?,?)",
                (cert_id, str(date.today()), datetime.now().strftime("%H:%M:%S"), acao, info)
            )
            conn.execute("""
                DELETE FROM historico WHERE cert_id=? AND id NOT IN (
                    SELECT id FROM historico WHERE cert_id=? ORDER BY id DESC LIMIT 90
                )
            """, (cert_id, cert_id))
    _executar_com_retry(_fazer)

def _migrar_colunas_log():
    """Adiciona colunas novas se o banco foi criado antes desta versão."""
    try:
        with _db_lock, _conectar() as conn:
            # log_emails
            colunas_log = [row[1] for row in conn.execute("PRAGMA table_info(log_emails)").fetchall()]
            if "lido" not in colunas_log:
                conn.execute("ALTER TABLE log_emails ADD COLUMN lido TEXT DEFAULT 'Pendente'")
            if "data_leitura" not in colunas_log:
                conn.execute("ALTER TABLE log_emails ADD COLUMN data_leitura TEXT DEFAULT ''")
            # certificados
            colunas_cert = [row[1] for row in conn.execute("PRAGMA table_info(certificados)").fetchall()]
            if "dias_alerta" not in colunas_cert:
                conn.execute("ALTER TABLE certificados ADD COLUMN dias_alerta INTEGER DEFAULT 15")
            if "enviar_alerta" not in colunas_cert:
                conn.execute("ALTER TABLE certificados ADD COLUMN enviar_alerta INTEGER DEFAULT 1")
    except Exception:
        pass

def registrar_log_email(cert: dict, destinatarios: list, assunto: str, status: str, erro: str = "", origem: str = "automatico"):
    def _fazer():
        with _db_lock, _conectar() as conn:
            conn.execute(
                "INSERT INTO log_emails (data_hora, cert_id, cert_nome, cert_tipo, destinatarios, assunto, status, erro, origem, lido, data_leitura) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    cert.get("id", ""),
                    cert.get("nome", ""),
                    cert.get("tipo", ""),
                    ", ".join(destinatarios),
                    assunto,
                    status,
                    erro,
                    origem,
                    "Pendente",
                    "",
                )
            )
    _executar_com_retry(_fazer)

def marcar_log_lido(log_id: int, lido: bool):
    """Marca um registro do log como lido ou pendente manualmente.
    Ao marcar como Lido, desativa automaticamente o lembrete do certificado.
    Ao marcar como Pendente, reativa o lembrete do certificado."""
    def _fazer():
        with _db_lock, _conectar() as conn:
            row = conn.execute("SELECT cert_id FROM log_emails WHERE id=?", (log_id,)).fetchone()
            cert_id = row["cert_id"] if row else None
            if lido:
                conn.execute(
                    "UPDATE log_emails SET lido='Lido', data_leitura=? WHERE id=?",
                    (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), log_id)
                )
                if cert_id:
                    conn.execute("UPDATE certificados SET enviar_alerta=0 WHERE id=?", (cert_id,))
            else:
                conn.execute(
                    "UPDATE log_emails SET lido='Pendente', data_leitura='' WHERE id=?",
                    (log_id,)
                )
                if cert_id:
                    conn.execute("UPDATE certificados SET enviar_alerta=1 WHERE id=?", (cert_id,))
    _executar_com_retry(_fazer)

def carregar_log_emails(filtro_data_ini="", filtro_data_fim="", filtro_cert="", filtro_status="", filtro_origem="", filtro_lido="") -> list:
    def _fazer():
        with _db_lock, _conectar() as conn:
            query = "SELECT * FROM log_emails WHERE 1=1"
            params = []
            if filtro_data_ini:
                query += " AND data_hora >= ?"
                params.append(filtro_data_ini + " 00:00:00")
            if filtro_data_fim:
                query += " AND data_hora <= ?"
                params.append(filtro_data_fim + " 23:59:59")
            if filtro_cert:
                query += " AND cert_nome LIKE ?"
                params.append(f"%{filtro_cert}%")
            if filtro_status and filtro_status != "Todos":
                query += " AND status = ?"
                params.append(filtro_status)
            if filtro_origem and filtro_origem != "Todos":
                query += " AND origem = ?"
                params.append(filtro_origem)
            if filtro_lido and filtro_lido != "Todos":
                query += " AND lido = ?"
                params.append(filtro_lido)
            query += " ORDER BY id DESC"
            rows = conn.execute(query, params).fetchall()
            return [dict(r) for r in rows]
    return _executar_com_retry(_fazer)

def atualizar_ultimo_alerta(cert_id: str, data: str):
    def _fazer():
        with _db_lock, _conectar() as conn:
            conn.execute("UPDATE certificados SET ultimo_alerta=? WHERE id=?", (data, cert_id))
    _executar_com_retry(_fazer)

# ─────────────────────────────────────────────
#  PERSISTENCIA — CONFIGURACOES
# ─────────────────────────────────────────────
def _get_config(chave: str, default=None):
    def _fazer():
        with _db_lock, _conectar() as conn:
            row = conn.execute("SELECT valor FROM configuracoes WHERE chave=?", (chave,)).fetchone()
            if row:
                try:
                    return json.loads(row["valor"])
                except Exception:
                    return row["valor"]
            return default
    return _executar_com_retry(_fazer)

def _get_config_original(chave: str, default=None):
    with _db_lock, _conectar() as conn:
        row = conn.execute("SELECT valor FROM configuracoes WHERE chave=?", (chave,)).fetchone()
        if row:
            try:
                return json.loads(row["valor"])
            except Exception:
                return row["valor"]
        return default

def _set_config(chave: str, valor):
    def _fazer():
      with _db_lock, _conectar() as conn:
        conn.execute(
            "INSERT INTO configuracoes (chave, valor) VALUES (?,?) "
            "ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor",
            (chave, json.dumps(valor, ensure_ascii=False))
          )
    _executar_com_retry(_fazer)

def carregar_config_email() -> dict:
    return _get_config("config_email", {})

def salvar_config_email(cfg: dict):
    _set_config("config_email", cfg)

def carregar_template() -> dict:
    return _get_config("template_email", TEMPLATE_PADRAO)

def salvar_template(t: dict):
    _set_config("template_email", t)

def salvar_template_verificado(t: dict) -> tuple:
    """Salva o template e confirma a gravacao lendo de volta do banco.
    Retorna (sucesso: bool, mensagem: str)."""
    try:
        _set_config("template_email", t)
    except Exception as e:
        return False, f"Erro ao gravar no banco de dados:\n{e}"
    try:
        salvo = _get_config("template_email", None)
    except Exception as e:
        return False, f"Nao foi possivel confirmar a gravacao:\n{e}"
    if not salvo or salvo.get("assunto") != t.get("assunto") or salvo.get("corpo") != t.get("corpo"):
        return False, "A gravacao nao foi confirmada no banco de dados.\nVerifique a conexao com a pasta de rede e tente novamente."
    return True, "Template salvo com sucesso!"

# ─────────────────────────────────────────────
#  LEITURA DE CERTIFICADO A1
# ─────────────────────────────────────────────
def ler_certificado_pfx(caminho, senha):
    with open(caminho, "rb") as f:
        dados = f.read()
    _, cert, _ = load_key_and_certificates(
        dados, senha.encode() if senha else None, default_backend()
    )
    nome = cert.subject.get_attributes_for_oid(
        __import__("cryptography.x509.oid", fromlist=["NameOID"]).NameOID.COMMON_NAME
    )[0].value
    vencimento = cert.not_valid_after_utc.date() if hasattr(cert, "not_valid_after_utc") \
                 else cert.not_valid_after.date()
    return nome, str(vencimento)

def ler_certificado_pem(caminho):
    with open(caminho, "rb") as f:
        dados = f.read()
    cert = load_pem_x509_certificate(dados, default_backend())
    nome = cert.subject.get_attributes_for_oid(
        __import__("cryptography.x509.oid", fromlist=["NameOID"]).NameOID.COMMON_NAME
    )[0].value
    vencimento = cert.not_valid_after_utc.date() if hasattr(cert, "not_valid_after_utc") \
                 else cert.not_valid_after.date()
    return nome, str(vencimento)

# ─────────────────────────────────────────────
#  E-MAIL
# ─────────────────────────────────────────────
def _cor_situacao(dias_restantes):
    if dias_restantes < 0:
        return "#c0392b", f"VENCIDO ha {abs(dias_restantes)} dia(s)"
    elif dias_restantes <= 7:
        return "#e74c3c", f"vence em {dias_restantes} dia(s) - CRITICO"
    elif dias_restantes <= 15:
        return "#e67e22", f"vence em {dias_restantes} dia(s)"
    else:
        return "#f39c12", f"vence em {dias_restantes} dia(s)"

def enviar_email(config, destinatarios, cert):
    if not config.get("smtp_host"):
        return False, "E-mail nao configurado."

    dias_restantes = (date.fromisoformat(cert["vencimento"]) - date.today()).days
    cor, situacao  = _cor_situacao(dias_restantes)
    template = carregar_template()

    variaveis = {
        "nome":        cert.get("nome", ""),
        "tipo":        cert.get("tipo", ""),
        "responsavel": cert.get("responsavel", "-"),
        "vencimento":  cert.get("vencimento", ""),
        "obs":         cert.get("obs", "-"),
        "dias":        str(dias_restantes),
        "situacao":    situacao,
        "cor":         cor,
    }

    try:
        assunto = template["assunto"].format(**variaveis)
        html    = template["corpo"].format(**variaveis)
    except KeyError as e:
        return False, f"Variavel invalida no template: {e}"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = assunto
    msg["From"]    = config["usuario"]
    msg["To"]      = ", ".join(destinatarios)
    # Solicita confirmação de leitura ao destinatário
    msg["Disposition-Notification-To"] = config["usuario"]
    msg["Return-Receipt-To"]           = config["usuario"]
    msg.attach(MIMEText(html, "html"))

    try:
        ctx   = ssl.create_default_context()
        porta = int(config.get("smtp_porta", 587))
        host  = config["smtp_host"]
        if porta == 465:
            with smtplib.SMTP_SSL(host, porta, context=ctx) as s:
                s.login(config["usuario"], config["senha"])
                s.sendmail(config["usuario"], destinatarios, msg.as_string())
        else:
            with smtplib.SMTP(host, porta, timeout=15) as s:
                s.ehlo()
                s.starttls(context=ctx)
                s.ehlo()
                s.login(config["usuario"], config["senha"])
                s.sendmail(config["usuario"], destinatarios, msg.as_string())
        return True, "E-mail enviado com sucesso."
    except Exception as e:
        return False, str(e)

# ─────────────────────────────────────────────
#  VERIFICACAO DIARIA
# ─────────────────────────────────────────────
def verificar_certificados(app=None):
    certs    = carregar_certificados()
    config   = carregar_config_email()
    hoje     = str(date.today())
    enviados = 0

    for cert in certs:
        try:
            dias = (date.fromisoformat(cert["vencimento"]) - date.today()).days
        except Exception:
            continue

        registrar_historico_db(cert["id"], "verificado", dias_restantes=dias)

        # Respeita configuração individual de envio de alerta
        if not cert.get("enviar_alerta", 1):
            continue

        dias_inicio = int(cert.get("dias_alerta") or 15)
        if dias > dias_inicio:
            continue
        if cert.get("ultimo_alerta") == hoje:
            continue

        destinatarios = [e.strip() for e in cert.get("emails", "").split(",") if e.strip()]
        if not destinatarios:
            destinatarios = [config.get("usuario", "")]
        destinatarios = [d for d in destinatarios if d]

        if destinatarios:
            ok, msg = enviar_email(config, destinatarios, cert)
            _tmpl = carregar_template()
            try:
                _assunto_log = _tmpl["assunto"].format(
                    nome=cert.get("nome",""), tipo=cert.get("tipo",""),
                    responsavel=cert.get("responsavel",""), vencimento=cert.get("vencimento",""),
                    obs=cert.get("obs",""), dias=str(dias), situacao="", cor=""
                )
            except Exception:
                _assunto_log = _tmpl.get("assunto", "")
            if ok:
                atualizar_ultimo_alerta(cert["id"], hoje)
                registrar_historico_db(cert["id"], "alerta_enviado",
                                       dias_restantes=dias, destinatarios=destinatarios)
                registrar_log_email(cert, destinatarios, _assunto_log, "Enviado", origem="automatico")
                enviados += 1
            else:
                registrar_historico_db(cert["id"], "erro_envio",
                                       dias_restantes=dias, erro=msg)
                registrar_log_email(cert, destinatarios, _assunto_log, "Erro", erro=msg, origem="automatico")

    if app:
        def _update():
            app.status_bar.config(
                text=f"  Ultima verificacao: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  {enviados} alerta(s) enviado(s)"
            )
            app.atualizar_tabela()
        app.after(0, _update)

def iniciar_scheduler(app):
    schedule.every().day.at("08:00").do(verificar_certificados, app=app)
    def loop():
        while True:
            schedule.run_pending()
            time.sleep(60)
    t = threading.Thread(target=loop, daemon=True)
    t.start()

# ═══════════════════════════════════════════════
#  JANELA: CONFIGURAR E-MAIL
# ═══════════════════════════════════════════════
class JanelaConfigEmail(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Configuracao de E-mail")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        self._carregar()

    def _build(self):
        hdr = tk.Frame(self, bg=COR_PRIMARIA, height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  Configuracao de E-mail SMTP",
                 bg=COR_PRIMARIA, fg="#ffffff",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=12, pady=10)

        pad   = {"padx": 10, "pady": 5}
        frame = ttk.LabelFrame(self, text="Parametros SMTP", padding=15)
        frame.pack(padx=20, pady=15, fill="x")

        campos = [
            ("Servidor SMTP:",    "smtp_host",  False),
            ("Porta SMTP:",       "smtp_porta", False),
            ("Usuario (e-mail):", "usuario",    False),
            ("Senha:",            "senha",      True),
        ]
        self.vars = {}
        for i, (label, key, pwd) in enumerate(campos):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky="w", **pad)
            v    = tk.StringVar()
            show = "*" if pwd else ""
            ttk.Entry(frame, textvariable=v, width=35, show=show).grid(row=i, column=1, **pad)
            self.vars[key] = v

        ttk.Label(frame,
            text="Gmail: smtp.gmail.com  |  Porta: 465  |  Use App Password",
            foreground="#888", font=("Arial", 8)
        ).grid(row=len(campos), column=0, columnspan=2, pady=(0, 5))

        bf = ttk.Frame(self)
        bf.pack(pady=(0, 15))
        ttk.Button(bf, text="Salvar",   command=self._salvar).pack(side="left", padx=5)
        ttk.Button(bf, text="Testar",   command=self._testar).pack(side="left", padx=5)
        ttk.Button(bf, text="Cancelar", command=self.destroy).pack(side="left", padx=5)

    def _carregar(self):
        cfg = carregar_config_email()
        for k, v in self.vars.items():
            v.set(cfg.get(k, ""))

    def _salvar(self):
        cfg = {k: v.get().strip() for k, v in self.vars.items()}
        salvar_config_email(cfg)
        messagebox.showinfo("Salvo", "Configuracao salva com sucesso!", parent=self)

    def _testar(self):
        cfg = {k: v.get().strip() for k, v in self.vars.items()}
        cert_teste = {
            "id": "teste", "nome": "Certificado de Teste", "tipo": "A1",
            "responsavel": "Teste", "vencimento": str(date.today()),
            "obs": "E-mail de teste do sistema"
        }
        ok, msg = enviar_email(cfg, [cfg.get("usuario", "")], cert_teste)
        if ok:
            messagebox.showinfo("Sucesso", "E-mail de teste enviado!", parent=self)
        else:
            messagebox.showerror("Erro", f"Falha ao enviar:\n{msg}", parent=self)

# ═══════════════════════════════════════════════
#  JANELA: CADASTRO / EDICAO DE CERTIFICADO
# ═══════════════════════════════════════════════
class JanelaCertificado(tk.Toplevel):
    def __init__(self, parent, cert=None, callback=None):
        super().__init__(parent)
        self.cert     = cert
        self.callback = callback
        self.title("Editar Certificado" if cert else "Novo Certificado")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        if cert:
            self._preencher()

    def _build(self):
        pad = {"padx": 10, "pady": 6}
        nb  = ttk.Notebook(self)
        nb.pack(padx=15, pady=10, fill="both")

        f_geral = ttk.Frame(nb, padding=10)
        nb.add(f_geral, text=" Dados Gerais ")

        ttk.Label(f_geral, text="Tipo:").grid(row=0, column=0, sticky="w", **pad)
        self.var_tipo = tk.StringVar(value="A1")
        ttk.Combobox(f_geral, textvariable=self.var_tipo,
                     values=["A1", "A3"], state="readonly", width=10
                     ).grid(row=0, column=1, sticky="w", **pad)

        ttk.Label(f_geral, text="Nome / Razao Social:").grid(row=1, column=0, sticky="w", **pad)
        self.var_nome = tk.StringVar()
        ttk.Entry(f_geral, textvariable=self.var_nome, width=38).grid(row=1, column=1, **pad)

        ttk.Label(f_geral, text="Responsavel:").grid(row=2, column=0, sticky="w", **pad)
        self.var_resp = tk.StringVar()
        ttk.Entry(f_geral, textvariable=self.var_resp, width=38).grid(row=2, column=1, **pad)

        ttk.Label(f_geral, text="Vencimento (AAAA-MM-DD):").grid(row=3, column=0, sticky="w", **pad)
        self.var_venc = tk.StringVar()
        ttk.Entry(f_geral, textvariable=self.var_venc, width=20).grid(row=3, column=1, sticky="w", **pad)

        ttk.Label(f_geral, text="Observacao:").grid(row=4, column=0, sticky="w", **pad)
        self.var_obs = tk.StringVar()
        ttk.Entry(f_geral, textvariable=self.var_obs, width=38).grid(row=4, column=1, **pad)

        ttk.Label(f_geral, text="E-mails p/ alerta\n(separe por virgula):").grid(row=5, column=0, sticky="w", **pad)
        self.var_emails = tk.StringVar()
        ttk.Entry(f_geral, textvariable=self.var_emails, width=38).grid(row=5, column=1, **pad)

        # --- Configurações de alerta ---
        f_alerta = ttk.LabelFrame(f_geral, text="Configurações de Lembrete", padding=(10, 6))
        f_alerta.grid(row=6, column=0, columnspan=2, sticky="ew", padx=10, pady=(8, 2))

        self.var_enviar_alerta = tk.BooleanVar(value=True)
        chk = ttk.Checkbutton(
            f_alerta,
            text="Enviar lembrete de vencimento por e-mail",
            variable=self.var_enviar_alerta,
            command=self._toggle_alerta
        )
        chk.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 6))

        self.lbl_dias_alerta = ttk.Label(f_alerta, text="Iniciar envio (dias antes):")
        self.lbl_dias_alerta.grid(row=1, column=0, sticky="w", padx=(0, 6))
        self.var_dias_alerta = tk.StringVar(value="15")
        self.spin_dias_alerta = ttk.Spinbox(
            f_alerta, textvariable=self.var_dias_alerta,
            from_=1, to=365, width=6, state="normal"
        )
        self.spin_dias_alerta.grid(row=1, column=1, sticky="w")
        ttk.Label(f_alerta, text="dia(s) antes do vencimento",
                  foreground="#64748b").grid(row=1, column=2, sticky="w", padx=6)

        self.f_a1 = ttk.Frame(nb, padding=10)
        nb.add(self.f_a1, text=" Arquivo A1 ")

        ttk.Label(self.f_a1, text="Arquivo (.pfx / .pem):").grid(row=0, column=0, sticky="w", **pad)
        self.var_arquivo = tk.StringVar()
        ttk.Entry(self.f_a1, textvariable=self.var_arquivo, width=32).grid(row=0, column=1, **pad)
        ttk.Button(self.f_a1, text="...", command=self._browse).grid(row=0, column=2, padx=2)

        ttk.Label(self.f_a1, text="Senha do arquivo:").grid(row=1, column=0, sticky="w", **pad)
        self.var_senha = tk.StringVar()
        f_senha = ttk.Frame(self.f_a1)
        f_senha.grid(row=1, column=1, sticky="w", **pad)
        self.entry_senha = ttk.Entry(f_senha, textvariable=self.var_senha, show="*", width=26)
        self.entry_senha.pack(side="left")
        self._senha_visivel = False
        ttk.Button(f_senha, text="Ver", width=4, command=self._toggle_senha).pack(side="left", padx=2)

        ttk.Label(self.f_a1,
                  text="A senha e armazenada de forma criptografada.\nO arquivo do certificado sera salvo internamente no banco de dados.",
                  foreground="#888", font=("Arial", 8)).grid(row=2, column=0, columnspan=3, sticky="w", padx=10)

        f_btns_a1 = ttk.Frame(self.f_a1)
        f_btns_a1.grid(row=3, column=0, columnspan=3, pady=10)
        ttk.Button(f_btns_a1, text="Ler certificado automaticamente",
                   command=self._ler_arquivo).pack(side="left", padx=4)
        self.btn_exportar = ttk.Button(f_btns_a1, text="Exportar arquivo salvo",
                   command=self._exportar_arquivo)
        self.btn_exportar.pack(side="left", padx=4)

        self.lbl_arquivo_salvo = ttk.Label(self.f_a1, text="", foreground="#16a34a",
                                            font=("Segoe UI", 8, "italic"))
        self.lbl_arquivo_salvo.grid(row=4, column=0, columnspan=3, sticky="w", padx=10)

        bf = ttk.Frame(self)
        bf.pack(pady=(0, 12))
        ttk.Button(bf, text="Salvar",   command=self._salvar).pack(side="left", padx=6)
        ttk.Button(bf, text="Cancelar", command=self.destroy).pack(side="left", padx=6)

    def _toggle_alerta(self):
        """Habilita/desabilita o campo de dias conforme o checkbox."""
        estado = "normal" if self.var_enviar_alerta.get() else "disabled"
        self.spin_dias_alerta.config(state=estado)
        self.lbl_dias_alerta.config(foreground="" if self.var_enviar_alerta.get() else "#94a3b8")

    def _toggle_senha(self):
        if not pedir_senha_mestre(self):
            return
        self._senha_visivel = not self._senha_visivel
        self.entry_senha.config(show="" if self._senha_visivel else "*")

    def _exportar_arquivo(self):
        b64 = getattr(self, "_arquivo_b64", None) or (
            self.cert.get("arquivo_b64") if self.cert else None
        )
        if not b64:
            messagebox.showwarning("Atencao", "Nenhum arquivo armazenado neste certificado.", parent=self)
            return
        ext = ".pfx" if self.cert and "pfx" in self.cert.get("arquivo_ext", "pfx") else ".pem"
        dest = filedialog.asksaveasfilename(
            defaultextension=ext,
            filetypes=[("Certificado", f"*{ext}"), ("Todos", "*.*")],
            initialfile=f"{self.cert.get('nome','certificado')}{ext}" if self.cert else f"certificado{ext}"
        )
        if dest:
            try:
                base64_para_arquivo(b64, dest)
                messagebox.showinfo("Sucesso", f"Arquivo exportado para:\n{dest}", parent=self)
            except Exception as e:
                messagebox.showerror("Erro", str(e), parent=self)

    def _browse(self):
        path = filedialog.askopenfilename(
            filetypes=[("Certificados", "*.pfx *.pem"), ("Todos", "*.*")]
        )
        if path:
            self.var_arquivo.set(path)
            try:
                self._arquivo_b64 = arquivo_para_base64(path)
                self._arquivo_ext = "pfx" if path.lower().endswith(".pfx") else "pem"
                self.lbl_arquivo_salvo.config(text="Arquivo carregado e sera salvo no banco ao confirmar.")
            except Exception as e:
                self._arquivo_b64 = None
                messagebox.showerror("Erro", f"Nao foi possivel ler o arquivo:\n{e}", parent=self)

    def _ler_arquivo(self):
        path      = self.var_arquivo.get().strip()
        senha_raw = self.var_senha.get().strip()
        senha     = descriptografar_senha(senha_raw) or senha_raw
        if not path:
            messagebox.showwarning("Atencao", "Selecione um arquivo primeiro.", parent=self)
            return
        try:
            if path.lower().endswith(".pfx"):
                nome, venc = ler_certificado_pfx(path, senha)
            else:
                nome, venc = ler_certificado_pem(path)
            self.var_nome.set(nome)
            self.var_venc.set(venc)
            messagebox.showinfo("Sucesso", f"Certificado lido!\nNome: {nome}\nVencimento: {venc}", parent=self)
        except Exception as e:
            messagebox.showerror("Erro", f"Nao foi possivel ler o arquivo:\n{e}", parent=self)

    def _preencher(self):
        c = self.cert
        self.var_tipo.set(c.get("tipo", "A1"))
        self.var_nome.set(c.get("nome", ""))
        self.var_resp.set(c.get("responsavel", ""))
        self.var_venc.set(c.get("vencimento", ""))
        self.var_obs.set(c.get("obs", ""))
        self.var_emails.set(c.get("emails", ""))
        enviar = bool(c.get("enviar_alerta", 1))
        self.var_enviar_alerta.set(enviar)
        self.var_dias_alerta.set(str(c.get("dias_alerta") or 15))
        self._toggle_alerta()
        self.var_arquivo.set(c.get("arquivo_nome", ""))
        # O campo de senha sempre inicia em branco por seguranca.
        # Se o usuario nao digitar nada, a senha ja salva e preservada.
        self.var_senha.set("")
        self._arquivo_b64 = c.get("arquivo_b64", None)
        self._arquivo_ext = c.get("arquivo_ext", "pfx")
        if self._arquivo_b64:
            self.lbl_arquivo_salvo.config(text="Arquivo ja armazenado no banco. Use 'Exportar' para recuperar.")

    def _salvar(self):
        nome = self.var_nome.get().strip()
        venc = self.var_venc.get().strip()
        if not nome or not venc:
            messagebox.showwarning("Atencao", "Nome e Vencimento sao obrigatorios.", parent=self)
            return
        try:
            date.fromisoformat(venc)
        except ValueError:
            messagebox.showerror("Erro", "Data invalida. Use o formato AAAA-MM-DD.", parent=self)
            return

        # Campo de senha sempre comeca em branco. Se o usuario nao digitar
        # nada durante uma edicao, mantem a senha ja armazenada intacta.
        senha_digitada = self.var_senha.get().strip()
        if senha_digitada:
            senha_enc = criptografar_senha(senha_digitada)
        elif self.cert:
            senha_enc = self.cert.get("senha_enc", "")
        else:
            senha_enc = ""

        cert_existente = self.cert or {}
        novo_b64  = getattr(self, "_arquivo_b64", None)
        novo_ext  = getattr(self, "_arquivo_ext", "pfx")
        b64_final = novo_b64 or cert_existente.get("arquivo_b64", "")
        ext_final = novo_ext if novo_b64 else cert_existente.get("arquivo_ext", "pfx")
        nome_arq  = os.path.basename(self.var_arquivo.get().strip()) if self.var_arquivo.get().strip() else cert_existente.get("arquivo_nome", "")

        cert_id = self.cert["id"] if self.cert else str(int(time.time()))
        is_novo = self.cert is None

        # Certificados novos sempre iniciam com o lembrete ativo automaticamente,
        # independente do estado do checkbox no momento do salvamento.
        # Alem disso, se o vencimento foi alterado (certificado renovado),
        # reativa automaticamente o lembrete, pois trata-se de um novo ciclo
        # de alerta (mesmo que o e-mail do ciclo anterior tenha sido marcado
        # como lido, o que desativa a flag).
        venc_anterior = cert_existente.get("vencimento", "")
        venc_mudou    = (not is_novo) and venc_anterior and venc_anterior != venc
        if is_novo or venc_mudou:
            enviar_alerta_final = True
        else:
            enviar_alerta_final = self.var_enviar_alerta.get()

        dados = {
            "id":            cert_id,
            "tipo":          self.var_tipo.get(),
            "nome":          nome,
            "responsavel":   self.var_resp.get().strip(),
            "vencimento":    venc,
            "obs":           self.var_obs.get().strip(),
            "emails":        self.var_emails.get().strip(),
            "arquivo_nome":  nome_arq,
            "arquivo_b64":   b64_final,
            "arquivo_ext":   ext_final,
            "senha_enc":     senha_enc,
            # Se o vencimento mudou (renovacao), zera o ultimo_alerta para que
            # o novo ciclo de lembretes comece do zero.
            "ultimo_alerta": "" if venc_mudou else cert_existente.get("ultimo_alerta", ""),
            "dias_alerta":   int(self.var_dias_alerta.get() or 15),
            "enviar_alerta": enviar_alerta_final,
        }

        salvar_certificado(dados)
        registrar_historico_db(cert_id, "cadastrado" if is_novo else "editado")

        if self.callback:
            self.callback()
        self.destroy()

# ═══════════════════════════════════════════════
#  JANELA PRINCIPAL
# ═══════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gerenciador de Certificados Digitais")
        self.geometry("1100x620")
        self.minsize(900, 480)
        self.configure(bg=COR_BG)
        self._tray_icon = None
        self._apply_style()
        self._build_header()
        self._build_toolbar()
        self._build_table()
        self._build_statusbar()
        self.atualizar_tabela()
        self._iniciar_auto_refresh()
        iniciar_scheduler(self)
        threading.Thread(target=verificar_certificados, args=(self,), daemon=True).start()
        if TRAY_DISPONIVEL:
            self.protocol("WM_DELETE_WINDOW", self._minimizar_tray)
            threading.Thread(target=self._iniciar_tray, daemon=True).start()
        else:
            self.protocol("WM_DELETE_WINDOW", self._confirmar_saida)

    def _criar_icone_tray(self):
        img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        draw.ellipse([2, 2, 62, 62], fill=(26, 42, 74, 255))
        draw.ellipse([14, 14, 50, 50], fill=(37, 99, 235, 255))
        draw.ellipse([20, 20, 44, 44], fill=(26, 42, 74, 255))
        draw.ellipse([40, 40, 56, 56], fill=(22, 163, 74, 255))
        return img

    def _iniciar_tray(self):
        icone = self._criar_icone_tray()
        menu = pystray.Menu(
            pystray.MenuItem("Abrir",             lambda: self.after(0, self._restaurar_janela), default=True),
            pystray.MenuItem("Verificar agora",   lambda: self.after(0, self.verificar_agora)),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("Iniciar com o Windows",
                             lambda: self.after(0, self.toggle_startup),
                             checked=lambda item: _startup_habilitado()),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("Sair",              lambda: self.after(0, self._sair_completo)),
        )
        self._tray_icon = pystray.Icon(
            "CertificadosDigitais", icone, "Gerenciador de Certificados", menu
        )
        self._tray_icon.run()

    def _minimizar_tray(self):
        self.withdraw()
        if self._tray_icon and TRAY_DISPONIVEL:
            self._tray_icon.visible = True

    def _restaurar_janela(self):
        self.deiconify()
        self.lift()
        self.focus_force()

    def _sair_completo(self):
        if self._tray_icon:
            self._tray_icon.stop()
        self.destroy()

    def _confirmar_saida(self):
        if messagebox.askyesno("Sair", "Deseja encerrar o Gerenciador de Certificados?"):
            self.destroy()

    def _apply_style(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".",
            background=COR_BG, foreground="#1e293b", font=("Segoe UI", 9))
        style.configure("Treeview",
            background=COR_BG_TABLE, fieldbackground=COR_BG_TABLE,
            foreground="#1e293b", rowheight=28, font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
            background=COR_PRIMARIA, foreground=COR_TEXTO_CLR,
            font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview",
            background=[("selected", COR_SECUNDARIA)],
            foreground=[("selected", "#ffffff")])

    def _build_header(self):
        hdr = tk.Frame(self, bg=COR_PRIMARIA, height=56)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr,
            text="  Gerenciador de Certificados Digitais",
            bg=COR_PRIMARIA, fg="#ffffff",
            font=("Segoe UI", 14, "bold")
        ).pack(side="left", padx=18, pady=10)
        self.lbl_clock = tk.Label(hdr, bg=COR_PRIMARIA, fg="#94a3b8",
                                   font=("Segoe UI", 9))
        self.lbl_clock.pack(side="right", padx=16)
        self._tick()
        self._build_menu()

    def _tick(self):
        self.lbl_clock.config(text=datetime.now().strftime("%d/%m/%Y   %H:%M:%S"))
        self.after(1000, self._tick)

    def _build_menu(self):
        mb = tk.Menu(self, bg=COR_PRIMARIA, fg="#ffffff",
                     activebackground=COR_SECUNDARIA, activeforeground="#ffffff",
                     borderwidth=0)
        self.config(menu=mb)
        m_cert = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="Certificados", menu=m_cert)
        m_cert.add_command(label="Novo certificado",    command=self.novo_cert)
        m_cert.add_command(label="Editar selecionado",  command=self.editar_cert)
        m_cert.add_command(label="Excluir selecionado", command=self.excluir_cert)
        m_cert.add_separator()
        m_cert.add_command(label="Sair", command=self.quit)
        m_conf = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="Configuracoes", menu=m_conf)
        m_conf.add_command(label="Configurar e-mail",  command=self.config_email)
        m_conf.add_command(label="Template do e-mail", command=self.config_template)
        m_conf.add_separator()
        m_conf.add_command(label="Iniciar com o Windows", command=self.toggle_startup)
        m_acao = tk.Menu(mb, tearoff=0)
        mb.add_cascade(label="Acoes", menu=m_acao)
        m_acao.add_command(label="Verificar agora", command=self.verificar_agora)
        m_acao.add_command(label="Atualizar lista", command=self.atualizar_tabela)
        m_acao.add_separator()
        m_acao.add_command(label="Log de E-mails",  command=self.abrir_log_emails)

    def _build_toolbar(self):
        tb = tk.Frame(self, bg="#223366", pady=6)
        tb.pack(fill="x")

        def btn(text, cmd, color=COR_SECUNDARIA):
            b = tk.Button(tb, text=text, command=cmd,
                          bg=color, fg="#ffffff",
                          activebackground="#1d4ed8", activeforeground="#ffffff",
                          relief="flat", font=("Segoe UI", 9, "bold"),
                          padx=10, pady=5, cursor="hand2", borderwidth=0)
            b.pack(side="left", padx=3)
            return b

        def sep():
            tk.Frame(tb, bg="#445577", width=1).pack(side="left", fill="y", padx=6, pady=4)

        btn("+ Novo",          self.novo_cert,       "#16a34a")
        btn("Editar",          self.editar_cert)
        btn("Excluir",         self.excluir_cert,    "#dc2626")
        btn("Historico",       self.ver_historico)
        sep()
        btn("↻ Atualizar",    self.atualizar_manual,  "#0f766e")
        btn("Verificar agora", self.verificar_agora,  "#0284c7")
        btn("Log E-mails",     self.abrir_log_emails,  "#b45309")
        btn("Config. E-mail",  self.config_email)
        btn("Template",        self.config_template)
        sep()

        tk.Label(tb, text="Filtro:", bg="#223366", fg="#cbd5e1",
                 font=("Segoe UI", 9)).pack(side="left")
        self.var_filtro = tk.StringVar()
        self.var_filtro.trace_add("write", lambda *_: self.atualizar_tabela())
        tk.Entry(tb, textvariable=self.var_filtro, width=20,
                 font=("Segoe UI", 9), relief="flat",
                 bg="#334466", fg="#ffffff", insertbackground="#ffffff"
                 ).pack(side="left", padx=(4, 2), ipady=3)

        self.var_tipo_filtro = tk.StringVar(value="Todos")
        cb = ttk.Combobox(tb, textvariable=self.var_tipo_filtro,
                          values=["Todos", "A1", "A3"], state="readonly", width=8)
        cb.pack(side="left", padx=2)
        self.var_tipo_filtro.trace_add("write", lambda *_: self.atualizar_tabela())

    def _build_table(self):
        cols    = ("tipo", "nome", "responsavel", "vencimento", "dias", "situacao", "emails")
        headers = ("Tipo", "Nome / Razao Social", "Responsavel", "Vencimento", "Dias", "Situacao", "E-mails")

        frame = tk.Frame(self, bg=COR_BG)
        frame.pack(fill="both", expand=True, padx=10, pady=(6, 0))

        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse")
        widths    = [50, 180, 120, 100, 50, 80, 160]
        stretches = {"nome": True, "emails": True}
        for col, hdr, w in zip(cols, headers, widths):
            self.tree.heading(col, text=hdr, command=lambda c=col: self._ordenar(c))
            anchor = "center" if col in ("tipo", "dias", "situacao", "vencimento") else "w"
            self.tree.column(col, width=w, minwidth=w, anchor=anchor, stretch=stretches.get(col, False))

        sb_y = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        sb_x = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        self.tree.tag_configure("vencido", background=COR_VENCIDO, foreground=COR_VENCIDO_FG)
        self.tree.tag_configure("critico", background=COR_CRITICO, foreground=COR_CRITICO_FG)
        self.tree.tag_configure("atencao", background=COR_ATENCAO, foreground=COR_ATENCAO_FG)
        self.tree.tag_configure("ok",      background=COR_OK,      foreground=COR_OK_FG)

        self.tree.bind("<Double-1>", lambda _: self.editar_cert())

        self._btn_mail_widgets  = {}
        self._btn_senha_widgets = {}
        self._frame_table = frame
        self.tree.bind("<Configure>",  lambda _: self.after(30, self._reposicionar_botoes))
        self.tree.bind("<MouseWheel>", lambda _: self.after(30, self._reposicionar_botoes))
        self.tree.bind("<Button-4>",   lambda _: self.after(30, self._reposicionar_botoes))
        self.tree.bind("<Button-5>",   lambda _: self.after(30, self._reposicionar_botoes))

    def _reposicionar_botoes(self):
        total_w = sum(self.tree.column(c, "width") for c in self.tree["columns"])
        x_mail  = total_w - 95
        x_senha = total_w - 185

        for iid, btn in list(self._btn_mail_widgets.items()):
            try:
                bbox = self.tree.bbox(iid)
            except Exception:
                btn.place_forget()
                continue
            if bbox:
                _, y, _, h = bbox
                btn.place(in_=self.tree, x=x_mail + 4, y=y + 2, width=84, height=h - 4)
            else:
                btn.place_forget()

        for iid, btn in list(self._btn_senha_widgets.items()):
            try:
                bbox = self.tree.bbox(iid)
            except Exception:
                btn.place_forget()
                continue
            if bbox:
                _, y, _, h = bbox
                btn.place(in_=self.tree, x=x_senha + 4, y=y + 2, width=82, height=h - 4)
            else:
                btn.place_forget()

    def _build_statusbar(self):
        sb = tk.Frame(self, bg=COR_PRIMARIA, height=26)
        sb.pack(fill="x", side="bottom")
        sb.pack_propagate(False)
        self.status_bar = tk.Label(sb, text="  Pronto.",
                                    bg=COR_PRIMARIA, fg="#94a3b8",
                                    font=("Segoe UI", 8), anchor="w")
        self.status_bar.pack(fill="x", padx=4)

    def atualizar_tabela(self):
        for btn in self._btn_mail_widgets.values():
            btn.destroy()
        self._btn_mail_widgets.clear()
        for btn in self._btn_senha_widgets.values():
            btn.destroy()
        self._btn_senha_widgets.clear()

        for row in self.tree.get_children():
            self.tree.delete(row)

        certs  = carregar_certificados()
        filtro = self.var_filtro.get().lower()
        tipo_f = self.var_tipo_filtro.get()

        for c in certs:
            if tipo_f != "Todos" and c.get("tipo") != tipo_f:
                continue
            if filtro and filtro not in c.get("nome", "").lower() \
                       and filtro not in c.get("responsavel", "").lower():
                continue
            try:
                dias = (date.fromisoformat(c["vencimento"]) - date.today()).days
            except Exception:
                dias = 0

            if dias < 0:
                sit, tag = "Vencido", "vencido"
            elif dias <= 7:
                sit, tag = "Critico", "critico"
            elif dias <= 30:
                sit, tag = "Atencao", "atencao"
            else:
                sit, tag = "OK",      "ok"

            self.tree.insert("", "end", iid=c["id"], tags=(tag,), values=(
                c.get("tipo", ""), c.get("nome", ""), c.get("responsavel", ""),
                c.get("vencimento", ""), dias, sit, c.get("emails", "")
            ))

            cid = c["id"]
            btn = tk.Button(
                self.tree,
                text="Enviar E-mail",
                font=("Segoe UI", 8, "bold"),
                bg="#2563eb", fg="#ffffff",
                activebackground="#1d4ed8", activeforeground="#ffffff",
                relief="flat", cursor="hand2", borderwidth=0,
                command=lambda id_=cid: self._enviar_email_manual(id_)
            )
            self._btn_mail_widgets[cid] = btn

            btn_s = tk.Button(
                self.tree,
                text="Ver Senha",
                font=("Segoe UI", 8, "bold"),
                bg="#7c3aed", fg="#ffffff",
                activebackground="#6d28d9", activeforeground="#ffffff",
                relief="flat", cursor="hand2", borderwidth=0,
                command=lambda id_=cid: self._ver_senha(id_)
            )
            self._btn_senha_widgets[cid] = btn_s

        self.after(50, self._reposicionar_botoes)
        total = len(self.tree.get_children())
        agora = datetime.now().strftime("%H:%M:%S")
        self.status_bar.config(text=f"  {total} certificado(s) exibido(s).   |   Atualizado: {agora}")

    def _ordenar(self, col):
        rows = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        rows.sort()
        for i, (_, k) in enumerate(rows):
            self.tree.move(k, "", i)
        self.after(50, self._reposicionar_botoes)

    def _cert_selecionado(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atencao", "Selecione um certificado na lista.")
            return None
        cid   = sel[0]
        certs = carregar_certificados()
        return next((c for c in certs if c["id"] == cid), None)

    def _ver_senha(self, cert_id):
        certs = carregar_certificados()
        cert  = next((c for c in certs if c["id"] == cert_id), None)
        if not cert:
            return

        senha_enc = cert.get("senha_enc", "")
        if not senha_enc:
            messagebox.showinfo("Senha", "Nenhuma senha cadastrada para este certificado.", parent=self)
            return

        if not pedir_senha_mestre(self):
            return

        senha = descriptografar_senha(senha_enc)
        if not senha:
            messagebox.showerror("Erro", "Nao foi possivel descriptografar a senha.", parent=self)
            return

        win = tk.Toplevel(self)
        win.title("Senha do Certificado")
        win.resizable(False, False)
        win.grab_set()
        win.configure(bg="#1e1b4b")

        hdr = tk.Frame(win, bg="#4c1d95", height=50)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  Senha do Certificado",
                 bg="#4c1d95", fg="#ffffff",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=12)

        tk.Label(win, text=cert.get("nome", ""),
                 bg="#1e1b4b", fg="#c4b5fd",
                 font=("Segoe UI", 10)).pack(pady=(14, 2))
        tk.Label(win,
                 text=f"Tipo: {cert.get('tipo', '')}   |   Vencimento: {cert.get('vencimento', '')}",
                 bg="#1e1b4b", fg="#7c3aed",
                 font=("Segoe UI", 8)).pack(pady=(0, 10))

        f_senha = tk.Frame(win, bg="#2e1065", bd=0, pady=12, padx=20)
        f_senha.pack(fill="x", padx=30, pady=6)

        tk.Label(f_senha, text="Senha:", bg="#2e1065", fg="#a78bfa",
                 font=("Segoe UI", 9)).pack(anchor="w")

        var_senha = tk.StringVar(value=senha)
        f_row = tk.Frame(f_senha, bg="#2e1065")
        f_row.pack(fill="x", pady=(4, 0))

        entry = tk.Entry(f_row, textvariable=var_senha, state="readonly",
                         font=("Courier New", 14, "bold"),
                         readonlybackground="#3b0764", fg="#f0abfc",
                         relief="flat", bd=6, justify="center")
        entry.pack(side="left", fill="x", expand=True, ipady=6)

        def _copiar():
            win.clipboard_clear()
            win.clipboard_append(senha)
            btn_copiar.config(text="Copiado!", bg="#16a34a")
            win.after(1500, lambda: btn_copiar.config(text="Copiar", bg="#7c3aed"))

        btn_copiar = tk.Button(f_row, text="Copiar", command=_copiar,
                               bg="#7c3aed", fg="#ffffff", relief="flat",
                               font=("Segoe UI", 9, "bold"), padx=10, cursor="hand2")
        btn_copiar.pack(side="left", padx=(6, 0))

        tk.Label(win,
                 text="Esta janela fechara automaticamente em 30 segundos.",
                 bg="#1e1b4b", fg="#6b7280",
                 font=("Segoe UI", 8)).pack(pady=(10, 4))

        self._countdown_label = tk.Label(win, text="30", bg="#1e1b4b", fg="#f87171",
                                          font=("Segoe UI", 22, "bold"))
        self._countdown_label.pack()

        tk.Button(win, text="Fechar", command=win.destroy,
                  bg="#6d28d9", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=20, pady=6,
                  cursor="hand2").pack(pady=14)

        win.geometry("380x320")

        def _tick(n):
            if not win.winfo_exists():
                return
            self._countdown_label.config(text=str(n))
            if n <= 0:
                win.destroy()
            else:
                win.after(1000, lambda: _tick(n - 1))

        _tick(30)

    def _enviar_email_manual(self, cert_id):
        certs  = carregar_certificados()
        config = carregar_config_email()
        cert   = next((c for c in certs if c["id"] == cert_id), None)
        if not cert:
            return

        destinatarios = [e.strip() for e in cert.get("emails", "").split(",") if e.strip()]
        if not destinatarios:
            destinatarios = [config.get("usuario", "")]
        destinatarios = [d for d in destinatarios if d]

        if not destinatarios:
            messagebox.showwarning("Atencao", "Nenhum e-mail configurado para este certificado.")
            return

        self.status_bar.config(text=f"  Enviando e-mail para {', '.join(destinatarios)}...")
        self.update_idletasks()

        def _enviar():
            ok, msg = enviar_email(config, destinatarios, cert)
            _tmpl = carregar_template()
            try:
                _dias_r = (date.fromisoformat(cert["vencimento"]) - date.today()).days
                _assunto_log = _tmpl["assunto"].format(
                    nome=cert.get("nome",""), tipo=cert.get("tipo",""),
                    responsavel=cert.get("responsavel",""), vencimento=cert.get("vencimento",""),
                    obs=cert.get("obs",""), dias=str(_dias_r), situacao="", cor=""
                )
            except Exception:
                _assunto_log = _tmpl.get("assunto", "")
            if ok:
                registrar_historico_db(cert_id, "alerta_enviado",
                                       manual=True, destinatarios=destinatarios)
                registrar_log_email(cert, destinatarios, _assunto_log, "Enviado", origem="manual")
                self.after(0, lambda: [
                    messagebox.showinfo("Sucesso", "E-mail enviado para:\n" + "\n".join(destinatarios)),
                    self.status_bar.config(text=f"  E-mail enviado manualmente - {datetime.now().strftime('%H:%M:%S')}"),
                    self.atualizar_tabela()
                ])
            else:
                registrar_log_email(cert, destinatarios, _assunto_log, "Erro", erro=msg, origem="manual")
                self.after(0, lambda: [
                    messagebox.showerror("Erro", f"Falha ao enviar e-mail:\n{msg}"),
                    self.status_bar.config(text="  Erro ao enviar e-mail.")
                ])

        threading.Thread(target=_enviar, daemon=True).start()

    def novo_cert(self):
        JanelaCertificado(self, callback=self.atualizar_tabela)

    def editar_cert(self):
        cert = self._cert_selecionado()
        if cert:
            JanelaCertificado(self, cert=cert, callback=self.atualizar_tabela)

    def excluir_cert(self):
        cert = self._cert_selecionado()
        if not cert:
            return
        if messagebox.askyesno("Confirmar", f"Excluir o certificado '{cert['nome']}'?"):
            excluir_certificado_db(cert["id"])
            self.atualizar_tabela()

    def ver_historico(self):
        cert = self._cert_selecionado()
        if cert:
            JanelaHistorico(self, cert)

    def config_email(self):
        JanelaConfigEmail(self)

    def config_template(self):
        JanelaTemplate(self)

    def abrir_log_emails(self):
        JanelaLogEmails(self)

    def atualizar_manual(self):
        self.status_bar.config(text="  Atualizando lista...")
        self.update_idletasks()
        self.atualizar_tabela()

    def verificar_agora(self):
        self.status_bar.config(text="  Verificando certificados...")
        threading.Thread(target=verificar_certificados, args=(self,), daemon=True).start()

    def _iniciar_auto_refresh(self, intervalo_ms=30000):
        """Atualiza a tabela automaticamente a cada 30 segundos."""
        self._auto_refresh_intervalo = intervalo_ms
        self._agendar_refresh()

    def _agendar_refresh(self):
        self._auto_refresh_job = self.after(
            self._auto_refresh_intervalo, self._executar_refresh
        )

    def _executar_refresh(self):
        self.atualizar_tabela()
        self._agendar_refresh()

    def toggle_startup(self):
        if _startup_habilitado():
            _desabilitar_startup()
            messagebox.showinfo(
                "Inicializacao",
                "O programa foi REMOVIDO da inicializacao automatica do Windows."
            )
        else:
            _habilitar_startup()
            messagebox.showinfo(
                "Inicializacao",
                "O programa foi adicionado a inicializacao automatica do Windows.\n"
                "Ele sera iniciado minimizado na bandeja ao ligar o PC."
            )

# ═══════════════════════════════════════════════
#  JANELA: HISTORICO DO CERTIFICADO
# ═══════════════════════════════════════════════
class JanelaHistorico(tk.Toplevel):
    ACOES = {
        "cadastrado":     ("[+]", "Cadastrado"),
        "editado":        ("[E]", "Editado"),
        "verificado":     ("[V]", "Verificado"),
        "alerta_enviado": ("[@]", "Alerta enviado"),
        "erro_envio":     ("[X]", "Erro no envio"),
    }

    def __init__(self, parent, cert):
        super().__init__(parent)
        self.title(f"Historico - {cert['nome']}")
        self.geometry("720x420")
        self.grab_set()
        self._build(cert)

    def _build(self, cert):
        hdr = tk.Frame(self, bg=COR_PRIMARIA, height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  {cert['nome']}   |   Tipo: {cert['tipo']}",
                 bg=COR_PRIMARIA, fg="#ffffff",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=12, pady=10)

        cols    = ("data", "hora", "acao", "info")
        headers = ("Data", "Hora", "Acao", "Informacoes")

        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=10, pady=(8, 4))

        tree   = ttk.Treeview(frame, columns=cols, show="headings")
        widths = [90, 70, 160, 360]
        for col, hdr_txt, w in zip(cols, headers, widths):
            tree.heading(col, text=hdr_txt)
            tree.column(col, width=w)

        sb = ttk.Scrollbar(frame, command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        historico = cert.get("historico", [])
        for entry in reversed(historico):
            icone, label = self.ACOES.get(entry.get("acao", ""), ("[?]", entry.get("acao", "")))
            extras = []
            if "dias_restantes" in entry:
                extras.append(f"{entry['dias_restantes']} dia(s) restante(s)")
            if entry.get("manual"):
                extras.append("Envio manual")
            if "destinatarios" in entry:
                extras.append(f"Para: {', '.join(entry['destinatarios'])}")
            if "erro" in entry:
                extras.append(f"Erro: {entry['erro']}")
            info = " | ".join(extras) if extras else "-"
            tree.insert("", "end", values=(
                entry.get("data", ""),
                entry.get("hora", ""),
                f"{icone} {label}",
                info
            ))

        if not historico:
            tree.insert("", "end", values=("", "", "Nenhum registro ainda.", ""))

        tk.Button(self, text="Fechar", command=self.destroy,
                  bg=COR_SECUNDARIA, fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9), padx=16, pady=5, cursor="hand2"
                  ).pack(pady=8)

# ═══════════════════════════════════════════════
#  JANELA: EDITOR DE TEMPLATE DE E-MAIL
# ═══════════════════════════════════════════════
class JanelaTemplate(tk.Toplevel):
    VARIAVEIS = [
        ("{nome}",        "Nome / Razao Social"),
        ("{tipo}",        "Tipo (A1 ou A3)"),
        ("{responsavel}", "Responsavel"),
        ("{vencimento}",  "Data de vencimento"),
        ("{dias}",        "Dias ate o vencimento"),
        ("{situacao}",    "Texto da situacao"),
        ("{obs}",         "Observacao"),
        ("{cor}",         "Cor HTML conforme urgencia"),
    ]

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Template do E-mail")
        self.geometry("800x650")
        self.minsize(700, 480)
        self.resizable(True, True)
        self.grab_set()
        self._build()
        self._carregar()

    def _build(self):
        hdr = tk.Frame(self, bg=COR_PRIMARIA, height=44)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  Editor de Template de E-mail",
                 bg=COR_PRIMARIA, fg="#ffffff",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=12, pady=10)

        # Botoes fixados na base da janela (empacotados ANTES do conteudo
        # expansivel, garantindo que fiquem sempre visiveis, mesmo se a
        # janela for redimensionada ou o conteudo acima crescer demais).
        bf = tk.Frame(self, bg="#e2e8f0")
        bf.pack(fill="x", side="bottom", pady=10)
        btn_inner = tk.Frame(bf, bg="#e2e8f0")
        btn_inner.pack()
        tk.Button(btn_inner, text="Salvar", command=self._salvar,
                  bg="#16a34a", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=14, pady=6,
                  cursor="hand2").pack(side="left", padx=5)
        tk.Button(btn_inner, text="Pre-visualizar", command=self._preview,
                  bg=COR_SECUNDARIA, fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=14, pady=6,
                  cursor="hand2").pack(side="left", padx=5)
        tk.Button(btn_inner, text="Restaurar padrao", command=self._restaurar,
                  bg="#b45309", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=14, pady=6,
                  cursor="hand2").pack(side="left", padx=5)
        tk.Button(btn_inner, text="Cancelar", command=self.destroy,
                  bg="#64748b", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=14, pady=6,
                  cursor="hand2").pack(side="left", padx=5)

        f_top = ttk.LabelFrame(self, text="Assunto", padding=8)
        f_top.pack(fill="x", padx=12, pady=(10, 4), side="top")
        self.var_assunto = tk.StringVar()
        ttk.Entry(f_top, textvariable=self.var_assunto, width=90).pack(fill="x")

        f_vars = ttk.LabelFrame(self, text="Variaveis - clique para inserir", padding=6)
        f_vars.pack(fill="x", padx=12, pady=4, side="bottom")
        for i, (var, desc) in enumerate(self.VARIAVEIS):
            col = i % 4
            row = i // 4
            ttk.Button(f_vars, text=var, width=14,
                       command=lambda v=var: self._inserir_variavel(v)
                       ).grid(row=row*2, column=col, padx=4, pady=2, sticky="w")
            ttk.Label(f_vars, text=desc, foreground="#666",
                      font=("Arial", 7)).grid(row=row*2+1, column=col, padx=4, sticky="w")

        f_corpo = ttk.LabelFrame(self, text="Corpo do E-mail (HTML)", padding=8)
        f_corpo.pack(fill="both", expand=True, padx=12, pady=4, side="top")
        self.txt_corpo = tk.Text(f_corpo, wrap="word", font=("Courier New", 9),
                                  undo=True, height=12)
        sb = ttk.Scrollbar(f_corpo, command=self.txt_corpo.yview)
        self.txt_corpo.configure(yscrollcommand=sb.set)
        self.txt_corpo.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    def _carregar(self):
        t = carregar_template()
        self.var_assunto.set(t.get("assunto", TEMPLATE_PADRAO["assunto"]))
        self.txt_corpo.delete("1.0", "end")
        self.txt_corpo.insert("1.0", t.get("corpo", TEMPLATE_PADRAO["corpo"]))

    def _inserir_variavel(self, var):
        self.txt_corpo.insert(tk.INSERT, var)
        self.txt_corpo.focus()

    def _salvar(self):
        t = {
            "assunto": self.var_assunto.get().strip(),
            "corpo":   self.txt_corpo.get("1.0", "end").strip()
        }
        if not t["assunto"] or not t["corpo"]:
            messagebox.showwarning("Atencao", "Assunto e corpo do e-mail nao podem ficar vazios.", parent=self)
            return
        ok, msg = salvar_template_verificado(t)
        if ok:
            messagebox.showinfo("Salvo", msg, parent=self)
            self._carregar()
        else:
            messagebox.showerror("Erro ao salvar", msg, parent=self)

    def _restaurar(self):
        if messagebox.askyesno("Confirmar", "Restaurar o template padrao?", parent=self):
            salvar_template(TEMPLATE_PADRAO)
            self._carregar()

    def _preview(self):
        corpo   = self.txt_corpo.get("1.0", "end").strip()
        assunto = self.var_assunto.get()
        exemplo = {
            "nome": "Empresa Exemplo Ltda", "tipo": "A1",
            "responsavel": "Joao da Silva",
            "vencimento": str(date.today()),
            "dias": "10", "situacao": "vence em 10 dia(s)",
            "obs": "Renovar com urgencia", "cor": "#e74c3c",
        }
        try:
            corpo_ex = corpo
            for k, v in exemplo.items():
                corpo_ex = corpo_ex.replace("{" + k + "}", v)
            assunto_ex = assunto
            for k, v in exemplo.items():
                assunto_ex = assunto_ex.replace("{" + k + "}", v)
        except Exception as e:
            messagebox.showerror("Erro", str(e), parent=self)
            return

        win = tk.Toplevel(self)
        win.title("Pre-visualizacao")
        win.geometry("700x480")
        tk.Label(win, text=f"Assunto: {assunto_ex}",
                 font=("Segoe UI", 10, "bold"), fg="#1e293b"
                 ).pack(anchor="w", padx=12, pady=(10, 2))
        ttk.Separator(win).pack(fill="x", padx=12, pady=4)

        txt = tk.Text(win, wrap="word", font=("Segoe UI", 9), state="normal", relief="flat")
        sb2 = ttk.Scrollbar(win, command=txt.yview)
        txt.configure(yscrollcommand=sb2.set)
        txt.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=(4, 12))
        sb2.pack(side="right", fill="y", pady=(4, 12), padx=(0, 4))
        txt.insert("1.0", corpo_ex)
        txt.config(state="disabled")

        tk.Button(win, text="Fechar", command=win.destroy,
                  bg=COR_SECUNDARIA, fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9), padx=14, pady=4, cursor="hand2"
                  ).pack(pady=(0, 10))

# ═══════════════════════════════════════════════
#  JANELA: LOG DE E-MAILS ENVIADOS
# ═══════════════════════════════════════════════
class JanelaLogEmails(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Log de E-mails Enviados")
        self.geometry("1060x580")
        self.grab_set()
        self._registros_cache = []
        self._build()
        self._carregar()

    def _build(self):
        hdr = tk.Frame(self, bg=COR_PRIMARIA, height=44)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="  Log de E-mails Enviados",
                 bg=COR_PRIMARIA, fg="#ffffff",
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=12, pady=10)

        # Filtros
        f_filtros = tk.Frame(self, bg="#f1f5f9", pady=8)
        f_filtros.pack(fill="x", padx=10, pady=(6, 0))

        tk.Label(f_filtros, text="De:", bg="#f1f5f9", font=("Segoe UI", 9)).grid(row=0, column=0, padx=(8,2), sticky="w")
        self.var_dt_ini = tk.StringVar()
        ttk.Entry(f_filtros, textvariable=self.var_dt_ini, width=12).grid(row=0, column=1, padx=2)

        tk.Label(f_filtros, text="Ate:", bg="#f1f5f9", font=("Segoe UI", 9)).grid(row=0, column=2, padx=(8,2), sticky="w")
        self.var_dt_fim = tk.StringVar()
        ttk.Entry(f_filtros, textvariable=self.var_dt_fim, width=12).grid(row=0, column=3, padx=2)

        tk.Label(f_filtros, text="Certificado:", bg="#f1f5f9", font=("Segoe UI", 9)).grid(row=0, column=4, padx=(12,2), sticky="w")
        self.var_cert = tk.StringVar()
        ttk.Entry(f_filtros, textvariable=self.var_cert, width=22).grid(row=0, column=5, padx=2)

        tk.Label(f_filtros, text="Status:", bg="#f1f5f9", font=("Segoe UI", 9)).grid(row=0, column=6, padx=(12,2), sticky="w")
        self.var_status = tk.StringVar(value="Todos")
        ttk.Combobox(f_filtros, textvariable=self.var_status,
                     values=["Todos", "Enviado", "Erro"], state="readonly", width=10
                     ).grid(row=0, column=7, padx=2)

        tk.Label(f_filtros, text="Origem:", bg="#f1f5f9", font=("Segoe UI", 9)).grid(row=0, column=8, padx=(12,2), sticky="w")
        self.var_origem = tk.StringVar(value="Todos")
        ttk.Combobox(f_filtros, textvariable=self.var_origem,
                     values=["Todos", "automatico", "manual"], state="readonly", width=12
                     ).grid(row=0, column=9, padx=2)

        tk.Label(f_filtros, text="Leitura:", bg="#f1f5f9", font=("Segoe UI", 9)).grid(row=0, column=10, padx=(12,2), sticky="w")
        self.var_lido = tk.StringVar(value="Todos")
        ttk.Combobox(f_filtros, textvariable=self.var_lido,
                     values=["Todos", "Pendente", "Lido"], state="readonly", width=10
                     ).grid(row=0, column=11, padx=2)

        tk.Button(f_filtros, text="Filtrar", command=self._carregar,
                  bg=COR_SECUNDARIA, fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=10, pady=3, cursor="hand2"
                  ).grid(row=0, column=10, padx=(10,2))
        tk.Button(f_filtros, text="Limpar", command=self._limpar_filtros,
                  bg="#64748b", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9), padx=8, pady=3, cursor="hand2"
                  ).grid(row=0, column=13, padx=2)

        tk.Label(f_filtros, text="(formato: AAAA-MM-DD)", bg="#f1f5f9",
                 fg="#94a3b8", font=("Segoe UI", 7)).grid(row=1, column=0, columnspan=4, sticky="w", padx=8)

        # Tabela
        cols    = ("data_hora", "cert_nome", "cert_tipo", "destinatarios", "assunto", "status", "origem", "lido", "data_leitura", "erro")
        headers = ("Data/Hora", "Certificado", "Tipo", "Destinatarios", "Assunto", "Status", "Origem", "Leitura", "Data Leitura", "Erro")
        widths  = [130, 150, 45, 170, 200, 65, 75, 75, 110, 110]

        frame = tk.Frame(self, bg=COR_BG)
        frame.pack(fill="both", expand=True, padx=10, pady=(6, 0))

        self.tree = ttk.Treeview(frame, columns=cols, show="headings")
        for col, hdr_txt, w in zip(cols, headers, widths):
            self.tree.heading(col, text=hdr_txt)
            self.tree.column(col, width=w, minwidth=50)

        self.tree.tag_configure("enviado", foreground="#166534", background="#dcfce7")
        self.tree.tag_configure("erro",    foreground="#991b1b", background="#fee2e2")
        self.tree.tag_configure("lido",    foreground="#1e40af", background="#dbeafe")
        self.tree.bind("<Double-1>", self._duplo_clique)

        sb_y = ttk.Scrollbar(frame, orient="vertical",   command=self.tree.yview)
        sb_x = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        # Rodape
        f_bottom = tk.Frame(self, bg=COR_BG, pady=6)
        f_bottom.pack(fill="x", padx=10)

        self.lbl_total = tk.Label(f_bottom, text="", bg=COR_BG,
                                   fg="#64748b", font=("Segoe UI", 9))
        self.lbl_total.pack(side="left")

        tk.Button(f_bottom, text="✓ Marcar como Lido",
                  command=self._marcar_lido,
                  bg="#1d4ed8", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=10, pady=4, cursor="hand2"
                  ).pack(side="right", padx=4)
        tk.Button(f_bottom, text="✗ Marcar como Pendente",
                  command=self._marcar_pendente,
                  bg="#64748b", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9), padx=10, pady=4, cursor="hand2"
                  ).pack(side="right", padx=4)
        tk.Button(f_bottom, text="Exportar para Excel (.xlsx)",
                  command=self._exportar_excel,
                  bg="#16a34a", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9, "bold"), padx=12, pady=4, cursor="hand2"
                  ).pack(side="right", padx=4)
        tk.Button(f_bottom, text="Fechar", command=self.destroy,
                  bg="#64748b", fg="#ffffff", relief="flat",
                  font=("Segoe UI", 9), padx=12, pady=4, cursor="hand2"
                  ).pack(side="right", padx=4)

    def _duplo_clique(self, event):
        """Duplo clique alterna entre Lido/Pendente."""
        sel = self.tree.selection()
        if not sel:
            return
        iid  = sel[0]
        vals = self.tree.item(iid, "values")
        # coluna 7 = lido
        status_lido = vals[7] if len(vals) > 7 else "Pendente"
        log_id = self._id_do_iid(iid)
        if log_id is None:
            return
        novo_lido = status_lido != "Lido"
        marcar_log_lido(log_id, novo_lido)
        self._carregar()
        self._notificar_app()

    def _id_do_iid(self, iid):
        """Recupera o id do banco a partir do iid do treeview."""
        idx = self.tree.index(iid)
        if 0 <= idx < len(self._registros_cache):
            return self._registros_cache[idx]["id"]
        return None

    def _marcar_lido(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atencao", "Selecione um registro na lista.", parent=self)
            return
        log_id = self._id_do_iid(sel[0])
        if log_id:
            marcar_log_lido(log_id, True)
            self._carregar()
            self._notificar_app()

    def _marcar_pendente(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Atencao", "Selecione um registro na lista.", parent=self)
            return
        log_id = self._id_do_iid(sel[0])
        if log_id:
            marcar_log_lido(log_id, False)
            self._carregar()
            self._notificar_app()

    def _notificar_app(self):
        """Atualiza a tabela principal caso a janela pai esteja disponivel,
        pois marcar leitura pode ter alterado a flag de lembrete do certificado."""
        try:
            if hasattr(self.master, "atualizar_tabela"):
                self.master.atualizar_tabela()
        except Exception:
            pass

    def _limpar_filtros(self):
        self.var_dt_ini.set("")
        self.var_dt_fim.set("")
        self.var_cert.set("")
        self.var_status.set("Todos")
        self.var_origem.set("Todos")
        self.var_lido.set("Todos")
        self._carregar()

    def _carregar(self):
        for row in self.tree.get_children():
            self.tree.delete(row)

        registros = carregar_log_emails(
            filtro_data_ini=self.var_dt_ini.get().strip(),
            filtro_data_fim=self.var_dt_fim.get().strip(),
            filtro_cert=self.var_cert.get().strip(),
            filtro_status=self.var_status.get(),
            filtro_origem=self.var_origem.get(),
            filtro_lido=self.var_lido.get(),
        )

        for r in registros:
            if r.get("lido") == "Lido":
                tag = "lido"
            elif r["status"] == "Enviado":
                tag = "enviado"
            else:
                tag = "erro"
            self.tree.insert("", "end", tags=(tag,), values=(
                r["data_hora"],
                r["cert_nome"],
                r["cert_tipo"],
                r["destinatarios"],
                r["assunto"],
                r["status"],
                r["origem"],
                r.get("lido") or "Pendente",
                r.get("data_leitura") or "",
                r.get("erro") or "",
            ))

        self.lbl_total.config(text=f"{len(registros)} registro(s) encontrado(s)")
        self._registros_cache = registros

    def _exportar_excel(self):
        if not self._registros_cache:
            messagebox.showwarning("Atencao", "Nenhum registro para exportar.", parent=self)
            return

        destino = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            initialfile=f"log_emails_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not destino:
            return

        try:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Log de E-mails"

                cabecalhos = ["Data/Hora", "Certificado", "Tipo", "Destinatarios",
                              "Assunto", "Status", "Origem", "Leitura", "Data Leitura", "Erro"]
                chaves     = ["data_hora", "cert_nome", "cert_tipo", "destinatarios",
                              "assunto", "status", "origem", "lido", "data_leitura", "erro"]

                hdr_fill = PatternFill("solid", fgColor="1A2A4A")
                hdr_font = Font(bold=True, color="FFFFFF")
                for col_idx, cab in enumerate(cabecalhos, 1):
                    cell = ws.cell(row=1, column=col_idx, value=cab)
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal="center")

                for row_idx, r in enumerate(self._registros_cache, 2):
                    for col_idx, chave in enumerate(chaves, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=r.get(chave, ""))
                        if r["status"] == "Erro":
                            cell.fill = PatternFill("solid", fgColor="FEE2E2")
                        elif r["status"] == "Enviado":
                            cell.fill = PatternFill("solid", fgColor="DCFCE7")

                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

                wb.save(destino)
                messagebox.showinfo("Exportado", f"Arquivo Excel salvo em:\n{destino}", parent=self)

            except ImportError:
                import csv
                destino_csv = destino.replace(".xlsx", ".csv")
                with open(destino_csv, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.DictWriter(f, fieldnames=[
                        "data_hora","cert_nome","cert_tipo","destinatarios",
                        "assunto","status","origem","erro"
                    ])
                    writer.writeheader()
                    writer.writerows(self._registros_cache)
                messagebox.showinfo(
                    "Exportado",
                    f"openpyxl nao instalado. Salvo como CSV:\n{destino_csv}\n\n"
                    "Para Excel, instale: pip install openpyxl",
                    parent=self
                )

        except Exception as e:
            messagebox.showerror("Erro", f"Nao foi possivel exportar:\n{e}", parent=self)


# ─────────────────────────────────────────────
#  PONTO DE ENTRADA
# ─────────────────────────────────────────────
if __name__ == "__main__":
    _init_db()
    _migrar_colunas_log()
    app = App()
    app.mainloop()
